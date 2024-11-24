import asyncio

from api.models.bigquery import BigQuery
from api.models.v2.gcp_cost_resource import GCPCostResource
from api.models.kubecost import KubecostReport
from api.utils.decorator import *
from api.models.v2.__constant import TECHFAMILIES, TECHFAMILY_MFI
from api.utils.v2.conversion import Conversion
from api.utils.v2.date import count_days_in_month
from api.utils.v2.idle_cost import get_idle_cost
from api.utils.v2.mail_context import MailContext
from api.utils.v2.mailer import Mailer
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

import shutil

from api.utils.v2.merge_project import merge_gcp_cost


async def monthly_report(gcp_data, kubecost_data, excel_file, cud_data, shared_cost, shared_cud_cost):
    template_path = "api/templates/v2/template.xlsx"

    rates = BigQuery.get_current_conversion_rate()

    rate = {
        "idr": rate["currency_conversion_rate"]
        for rate in rates
    }

    # Styling Cells
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    fill_color = PatternFill(fgColor="b6d7a8", fill_type="solid")
    secondary_fill_color = PatternFill(fgColor="9fc5e8", fill_type="solid")

    if not os.path.exists(excel_file):
        # Copy the file
        shutil.copy(template_path, excel_file)

    workbook = load_workbook(excel_file)
    worksheet = workbook.get_sheet_by_name(gcp_data['__summary']['name'])

    services = {}

    for service, costs in gcp_data.items():
        if service == "__summary":
            continue
        total_current_cost = round(sum(cost["current_cost"] for cost in costs.values()), 2)
        if total_current_cost > 0:
            services[service] = total_current_cost

    total_current_cost = gcp_data["__summary"]['current_cost']
    total_current_cost_kubecost = kubecost_data["data"]["summary"]["cost_this_period"]
    positive_cud_cost = abs(cud_data)
    positive_shared_cud_cost = abs(shared_cud_cost)
    total_cud = positive_cud_cost + positive_shared_cud_cost
    total_after_cud = (total_current_cost + shared_cost) - total_cud

    kubecost_services = {}

    for service in kubecost_data["data"]["services"]:
        index_kubecost = (service["cost_this_period"] / total_current_cost_kubecost) * 100

        usd_cost = service["cost_this_period"]
        idr_app_cost = total_after_cud * (index_kubecost / 100)

        kubecost_services[service["service_name"]] = {
            "kubecost_usd": Conversion.usd_format(usd_cost),
            "kubecost_idr": Conversion.convert_idr(usd_cost, rate["idr"]),
            "app_cost_usd": Conversion.convert_usd(idr_app_cost, rate["idr"]),
            "app_cost_idr": Conversion.idr_format(idr_app_cost),
            "index_weight": index_kubecost
        }

    # Write GCP header row
    header = ['No', 'GCP Service Name', 'Total Cost (IDR)']
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.border = border
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.fill = fill_color

    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 50
    worksheet.column_dimensions['C'].width = 17

    worksheet.column_dimensions['D'].width = 15

    worksheet.column_dimensions['E'].width = 5
    worksheet.column_dimensions['F'].width = 35
    worksheet.column_dimensions['G'].width = 20
    worksheet.column_dimensions['H'].width = 20

    sorted_services = dict(sorted(services.items(), key=lambda item: item[1], reverse=True))

    # Write GCP data rows
    for row_num, service in enumerate(sorted_services, 1):

        cell1 = worksheet.cell(row=row_num + 1, column=1)
        cell2 = worksheet.cell(row=row_num + 1, column=2)
        cell3 = worksheet.cell(row=row_num + 1, column=3)

        cell1.value = row_num
        cell2.value = service
        cell3.value = Conversion.idr_format(sorted_services[service])

        cell1.border = border
        cell2.border = border
        cell3.border = border

        cell1.alignment = center_alignment

    # Write GCP TOTAL Cost
    total_rows = len(services) + 2
    worksheet.merge_cells(f"A{total_rows}:B{total_rows}")
    worksheet[f"A{total_rows}"].value = "Total"
    worksheet[f"A{total_rows}"].font = bold_font
    worksheet[f"A{total_rows}"].border = border
    worksheet[f"A{total_rows}"].fill = fill_color
    worksheet[f"C{total_rows}"].value = Conversion.idr_format(total_current_cost)
    worksheet[f"C{total_rows}"].font = bold_font
    worksheet[f"C{total_rows}"].border = border
    worksheet[f"C{total_rows}"].fill = fill_color

    # Write TOTAL SHARED COST
    worksheet.merge_cells(f"A{total_rows + 1}:B{total_rows + 1}")
    worksheet[f"A{total_rows + 1}"].value = "+Shared Cost (Infra, Data, ITCorp, Security & YBQTest)"
    worksheet[f"A{total_rows + 1}"].font = bold_font
    worksheet[f"A{total_rows + 1}"].border = border
    worksheet[f"A{total_rows + 1}"].fill = fill_color
    worksheet[f"C{total_rows + 1}"].value = Conversion.idr_format(shared_cost)
    worksheet[f"C{total_rows + 1}"].font = bold_font
    worksheet[f"C{total_rows + 1}"].border = border
    worksheet[f"C{total_rows + 1}"].fill = fill_color

    # Write CUD Total Cost
    worksheet.merge_cells(f"A{total_rows + 2}:B{total_rows + 2}")
    worksheet[f"A{total_rows + 2}"].value = "-CUD Cost"
    worksheet[f"A{total_rows + 2}"].font = bold_font
    worksheet[f"A{total_rows + 2}"].border = border
    worksheet[f"C{total_rows + 2}"].value = Conversion.idr_format(total_cud)
    worksheet[f"C{total_rows + 2}"].font = bold_font
    worksheet[f"C{total_rows + 2}"].border = border

    worksheet.merge_cells(f"A{total_rows + 3}:B{total_rows + 3}")
    worksheet[f"A{total_rows + 3}"].value = "TOTAL Cost After CUD"
    worksheet[f"A{total_rows + 3}"].font = bold_font
    worksheet[f"A{total_rows + 3}"].border = border
    worksheet[f"A{total_rows + 3}"].fill = fill_color
    worksheet[f"C{total_rows + 3}"].value = Conversion.idr_format(total_after_cud)
    worksheet[f"C{total_rows + 3}"].font = bold_font
    worksheet[f"C{total_rows + 3}"].border = border
    worksheet[f"C{total_rows + 3}"].fill = fill_color

    # Conversion rate row
    worksheet.merge_cells(f"A{total_rows + 5}:B{total_rows + 5}")
    worksheet[f"A{total_rows + 5}"].value = "Current GCP Conversion Rate"
    worksheet[f"A{total_rows + 5}"].font = bold_font
    worksheet[f"A{total_rows + 5}"].border = border
    worksheet[f"A{total_rows + 5}"].fill = secondary_fill_color
    worksheet[f"C{total_rows + 5}"].value = Conversion.idr_format(rate["idr"])
    worksheet[f"C{total_rows + 5}"].font = bold_font
    worksheet[f"C{total_rows + 5}"].border = border
    worksheet[f"C{total_rows + 5}"].fill = secondary_fill_color

    # Write Kubecost header row
    header = ['No', 'Service Name', 'Total Cost (USD)', 'Total Cost (IDR)']
    for col_num, column_title in enumerate(header, 5):
        for row in range(2):
            cell = worksheet.cell(row=row+1, column=col_num)
            cell.value = column_title
            cell.border = border
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.fill = fill_color

    worksheet.merge_cells("E1:E2")
    worksheet.merge_cells("F1:F2")
    worksheet.merge_cells("G1:H1")

    cell_g = worksheet.cell(row=1, column=7)
    cell_g.value = "Application Based Costs"

    sorted_kubecost = dict(sorted(kubecost_services.items(), key=lambda x: float(x[1]['app_cost_idr'][2:]
                                                                                 .replace('.', '')
                                                                                 .replace(',', '.')), reverse=True))
    # Write Kubecost rows data
    for row_num, service in enumerate(sorted_kubecost, 1):
        cell1 = worksheet.cell(row=row_num + 2, column=5)
        cell2 = worksheet.cell(row=row_num + 2, column=6)
        cell3 = worksheet.cell(row=row_num + 2, column=7)
        cell4 = worksheet.cell(row=row_num + 2, column=8)

        cell1.value = row_num
        cell2.value = service
        cell3.value = sorted_kubecost[service]["app_cost_usd"]
        cell4.value = sorted_kubecost[service]["app_cost_idr"]

        cell1.border = border
        cell2.border = border
        cell3.border = border
        cell4.border = border

        cell1.alignment = center_alignment

    # Write Kubecost TOTAL Cost
    total_rows = len(kubecost_services) + 3
    worksheet.merge_cells(f"E{total_rows}:F{total_rows}")
    worksheet[f"E{total_rows}"].value = "TOTAL"
    worksheet[f"E{total_rows}"].font = bold_font
    worksheet[f"E{total_rows}"].border = border
    worksheet[f"E{total_rows}"].fill = fill_color

    worksheet[f"G{total_rows}"].value = Conversion.convert_usd(total_after_cud, rate["idr"])
    worksheet[f"G{total_rows}"].font = bold_font
    worksheet[f"G{total_rows}"].border = border
    worksheet[f"G{total_rows}"].fill = fill_color

    worksheet[f"H{total_rows}"].value = Conversion.idr_format(total_after_cud)
    worksheet[f"H{total_rows}"].font = bold_font
    worksheet[f"H{total_rows}"].border = border
    worksheet[f"H{total_rows}"].fill = fill_color

    workbook.save(excel_file)


# Function to process the report generation and email sending
async def process_report(tech_family, kubecost_data, gcp_data, idle_data, index_weight,
                         conversion_rate, metadata, period, excel_file, cud_cost, shared_cost, shared_cud_cost):
    context = {}
    loop = asyncio.get_event_loop()

    # Distribute Idle cost for each tech family
    idle_cost_context = get_idle_cost(
        idle_data=idle_data,
        index_weight=index_weight
    )

    if idle_cost_context:
        context.update(idle_cost_context)

    if period == "monthly":
        await monthly_report(gcp_data, kubecost_data, excel_file, cud_cost, shared_cost, shared_cud_cost)
    else:
        # Create async tasks for generating GCP and Kubecost contexts
        async_tasks = [
            loop.run_in_executor(
                None, MailContext.gcp_context, gcp_data, conversion_rate
            ),
            loop.run_in_executor(None, MailContext.kubecost_context, kubecost_data),
        ]

        async_context = await asyncio.gather(*async_tasks)
        gcp_context = async_context[0]
        kubecost_context = async_context[1]

        context.update(await gcp_context)
        context.update(await kubecost_context)

        # Preparing data for email context
        date_time = kubecost_data["data"]["date"]
        to_email = kubecost_data["pic_email"]
        template_path = "v2/email_template.html"
        em_name = kubecost_data["pic"]
        subject = f"!!! Hi {em_name}, your GCP Cost on {date_time} !!!"
        tech_family_name = kubecost_data["tech_family"]
        project_name = f"({tech_family_name} - {kubecost_data['project']})"

        context["em_name"] = em_name
        context["project_name"] = project_name

        # Send the email report
        await asyncio.create_task(
            Mailer.send_report(
                to_email=to_email,
                subject=subject,
                template_path=template_path,
                context=context,
                em_name=em_name,
                tech_family=tech_family,
                metadata=metadata
            )
        )


# Decorators for validating request parameters
@async_date_validator
@async_period_validator
@async_user_validator
async def create_report(request, date=None, period=None):
    date = request.GET.get("date")
    period = request.GET.get("period")
    metadata = request.META,
    loop = asyncio.get_event_loop()

    # Determine the report period duration
    day = 7 if period == "weekly" else count_days_in_month(date) if period == "monthly" else 1

    # Create async tasks for fetching GCP and Kubecost data
    async_tasks = [
        loop.run_in_executor(None, GCPCostResource.get_cost, date, day),
        loop.run_in_executor(None, KubecostReport.report, date, period),
        loop.run_in_executor(None, GCPCostResource.get_cud_cost, date, day),
        loop.run_in_executor(None, GCPCostResource.get_shared_cost, date, day)
    ]

    result = await asyncio.gather(*async_tasks)

    gcp_result = result[0]
    kubecost_result = result[1]
    cud_result = result[2]
    shared_cost = result[3]

    # merged_gcp_result = merge_gcp_cost(gcp_result)

    # Extract idle data from Kubecost result
    idle_data = next(
        (
            service
            for service in kubecost_result["UNREGISTERED"]["data"]["services"]
            if service["service_name"] == "__idle__"
        ),
        None,
    )

    report_task = []

    excel_file = f"report_{date}.xlsx"

    # Generate reports for each tech family
    for techfamily in TECHFAMILIES:
        billing = "procar" if techfamily in TECHFAMILY_MFI else "moladin"
        project = "MFI" if techfamily in TECHFAMILY_MFI else "MDI"

        report_task.append(
            asyncio.create_task(
                process_report(
                    tech_family=techfamily,
                    kubecost_data=kubecost_result[techfamily],
                    gcp_data=gcp_result[billing][techfamily],
                    idle_data=idle_data["data"],
                    index_weight=gcp_result["__extras__"]["index_weight"][project][techfamily],
                    conversion_rate=gcp_result["__extras__"]["conversion_rate"],
                    metadata=metadata,
                    period=period,
                    excel_file=excel_file,
                    cud_cost=cud_result[project][techfamily],
                    shared_cost=shared_cost['shared_cost'][project][techfamily],
                    shared_cud_cost=shared_cost['cud_cost'][project][techfamily]
                )
            )
        )

    await asyncio.gather(*report_task)

    # Send monthly report just for VP
    if period == "monthly":
        period_date = gcp_result["moladin"]["__summary"]["usage_date_current"]

        subject = f"Monthly GCP Cost Report - {period_date}"
        template_path = "v2/monthly_email_template.html"
        context = {}
        await asyncio.create_task(
            Mailer.send_monthly_report(
                subject=subject,
                template_path=template_path,
                context=context,
                excel_file=excel_file
            )
        )

        os.remove(excel_file)

    return JsonResponse(
        {"success": True, "message": f"Report {period} email sent!", "data": "ok"}, status=200
    )
